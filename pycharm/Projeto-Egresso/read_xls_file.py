from openpyxl import load_workbook
from dataBaseConnect import ConnecyMyDataBase
import time

inicio = time.time()
print("Lendo o arquivo")
wb = load_workbook('pathfile', read_only=True)
print("Lendo as folhas")
for sheet_name in wb.sheetnames:
    print(sheet_name)
sheet = wb['Escolas_do_Ceara']

data = ""
first_line = ""
database = ConnecyMyDataBase()
print('Conectando com o banco de dados')
database.set_connection(user, password, host, database=schema)
print('Conexao com sucesso')
linha = 1
coluna = 1
# for i in range(1, sheet.max_row, 1):
# for j in range(0, sheet.max_column, 1):
for rows in sheet.iter_rows(min_row=2):

    for cell in rows:
        # tmp = sheet.cell(row=i+1, column=j + 1).value

        tmp = cell.value
        if tmp is None:
            tmp = "''"
        elif type(tmp) is str:
            tmp = "'" + tmp + "'"
        elif type(tmp) is int or type(tmp) is float:
            # tmp = str(int(sheet.cell(row=i+1, column=j + 1).value))
            tmp = "'" + str(int(tmp)) + "'"
        else:
            # tmp = str(sheet.cell(row=i+1, column=j + 1).value)
            tmp = "'" + tmp + "'"
        if coluna == 132:
            # first_line = first_line + sheet.cell(row=1, column=j + 1).value
            data = data + tmp
        else:
            # first_line = first_line + sheet.cell(row=1, column=j + 1).value + ","
            data = data + tmp + ","
            coluna = coluna + 1

    database.insert_table_many_data(data, linha)
    data = ''
    linha = linha + 1
    coluna = 1

database.get_connection().close()
database.get_connection_cursor().close()
fim = time.time()
total = fim - inicio
print(str(total))
